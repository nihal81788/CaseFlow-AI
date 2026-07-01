import os
from openpyxl import Workbook, load_workbook

EXCEL_FILE_PATH = os.path.join("data", "extracted_data.xlsx")

# Expected field order based on JSON output
COLUMNS = [
    "document_type",
    "case_number",
    "court_name",
    "judge",
    "hearing_date",
    "required_documents",
    "next_action",
    "summary"
]

def append_to_excel(data):
    """
    Appends the extracted JSON data to an Excel file as a new row.
    Ensures the columns are in the exact order specified by COLUMNS.
    """
    print("[ExcelExport] Appending row to Excel...")
    
    # Ensure data folder exists
    os.makedirs(os.path.dirname(EXCEL_FILE_PATH), exist_ok=True)
    
    # Check if the file already exists
    if not os.path.exists(EXCEL_FILE_PATH):
        # Create a new workbook and add headers
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Cases"
        ws.append(COLUMNS)
    else:
        # Load existing workbook
        try:
            wb = load_workbook(EXCEL_FILE_PATH)
            ws = wb.active
        except Exception as e:
            print(f"[ExcelExport] Failed to open existing Excel file: {e}")
            return
            
    # Prepare row data in the exact order of COLUMNS
    row = []
    for col in COLUMNS:
        row.append(str(data.get(col, "")))
        
    # Append the row
    ws.append(row)
    
    # Save the workbook
    try:
        wb.save(EXCEL_FILE_PATH)
        print(f"[ExcelExport] Successfully appended data to {EXCEL_FILE_PATH}")
    except Exception as e:
        print(f"[ExcelExport] Error saving Excel file. It might be open in another program: {e}")
